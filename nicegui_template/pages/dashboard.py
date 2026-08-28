"""
Dashboard v3 — dropdowns separate from content (stable), compare works.
"""

from datetime import datetime, timedelta
from html import unescape
import json
import os
import re
import tempfile
import time
from nicegui import ui
import pandas as pd
from services.dashboard_adapter import get_adapter

CARD = "background:#1e1e2e;border-radius:12px;padding:20px;box-shadow:0 4px 12px rgba(0,0,0,0.3);"
MV = "font-size:1.3rem;font-weight:700;color:#cdd6f4;"
ML = "font-size:0.8rem;color:#a6adc8;text-transform:uppercase;letter-spacing:0.5px;"
ST = "font-size:1.1rem;font-weight:600;color:#cdd6f4;margin-bottom:12px;"
SC = {
    "Keeper": "#10b981",
    "Optimasi": "#f59e0b",
    "Relocate": "#ef4444",
    "Tidak Aktif": "#94a3b8",
}

_daily_data_cache = None
_daily_data_cache_mtime = None


def _load_daily_data():
    """List of daily_summary records (cached by file mtime). [] kalau gagal."""
    global _daily_data_cache, _daily_data_cache_mtime
    try:
        path = os.path.abspath(
            os.path.join(
                os.path.dirname(__file__),
                "..", "..",
                "streamlit_template", "data", "api_cache", "daily_summary.json",
            )
        )
        mtime = os.path.getmtime(path)
        if _daily_data_cache is not None and _daily_data_cache_mtime == mtime:
            return _daily_data_cache
        with open(path) as f:
            data = json.load(f)
        _daily_data_cache = data
        _daily_data_cache_mtime = mtime
        return data
    except Exception:
        return []


def _load_last_tx_dates():
    """Map outlet_name -> tanggal transaksi terakhir (YYYY-MM-DD). {} kalau gagal."""
    last = {}
    for r in _load_daily_data():
        name = str(r.get("outlet_name", "")).strip()
        d = str(r.get("date", ""))[:10]
        if name and d > last.get(name, ""):
            last[name] = d
    return last


ECHART = {
    "backgroundColor": "#1e1e2e",
    "textStyle": {"color": "#cdd6f4"},
    "title": {"textStyle": {"color": "#cdd6f4"}},
    "legend": {"textStyle": {"color": "#a6adc8"}},
    "xAxis": {
        "axisLabel": {"color": "#a6adc8"},
        "axisLine": {"lineStyle": {"color": "#45475a"}},
    },
    "yAxis": {
        "axisLabel": {"color": "#a6adc8"},
        "splitLine": {"lineStyle": {"color": "#313244"}},
    },
}
TBL_CSS = """<style>
.tbl-gr{color:#a6e3a1!important;font-weight:600}
.tbl-rd{color:#f38ba8!important;font-weight:600}
.tbl-gd{color:#f9e2af!important;font-weight:600}
::-webkit-scrollbar{height:8px;width:8px}
::-webkit-scrollbar-track{background:#11111b;border-radius:4px}
::-webkit-scrollbar-thumb{background:#45475a;border-radius:4px}
::-webkit-scrollbar-thumb:hover{background:#585b70}
.ag-theme-balham-dark{--ag-background-color:#1e1e2e;--ag-header-background-color:#181825;--ag-odd-row-background-color:#1a1a2e;--ag-row-hover-color:#313244;--ag-border-color:#313244;--ag-font-size:12px;--ag-header-height:40px;--ag-row-height:40px;--ag-selected-row-background-color:#2a2a4e}
</style>"""

_df = None
_ff = None
_cp = None
_cmp = None
_content = None
_act_sel = None
_cmp_sel = None
_periods = []
_a = None


def _n(n):
    try:
        return f"{int(round(float(n))):,}".replace(",", ".")
    except:
        return str(n)


def _c(n):
    try:
        return f"Rp {int(round(float(n))):,}".replace(",", ".")
    except:
        return "Rp 0"


def _p(v):
    try:
        return f"{float(v):.1f}".replace(".", ",") + "%"
    except:
        return "0,0%"


def _build_outlet_table(cpv, cmv):
    """Rebuild just the outlet table content."""
    global _a
    a = _a
    src = _ff
    if _content is None:
        return
    _content.clear()
    if src is None or src.empty:
        return
    s = src.copy()
    for cl in [
        "total_revenue",
        "foto_qty",
        "unlock_qty",
        "print_qty",
        "paid_per_photo_rate",
    ]:
        if cl in s.columns:
            s[cl] = pd.to_numeric(s[cl], errors="coerce").fillna(0)
    if _cp and "periode" in s.columns:
        md = s[s["periode"].astype(str) == str(_cp)].copy()
    else:
        md = s.copy()

    with _content:
        mt = a.calculate_metrics(md)
        with ui.row().classes("w-full gap-4 mb-6"):
            for lb, vl, cl in [
                ("💰 Revenue", a.format_currency(mt["total_revenue"]), "#89b4fa"),
                ("🏪 Outlets", _n(mt["total_outlets"]), "#a6e3a1"),
                ("📈 Avg Conv Rate", f"{mt['avg_conversion']:.1f}%", "#f9e2af"),
                ("📸 Photos", _n(mt["total_photos"]), "#f38ba8"),
            ]:
                with ui.card().classes("flex-1 min-w-[150px]").style(CARD):
                    ui.label(lb).style(ML)
                    ui.label(vl).style(MV)
                    ui.label(f"Periode: {_cp or 'Semua'}").classes(
                        "text-[10px] text-gray-600 mt-1"
                    )
        ui.separator().classes("mb-4")
        ui.label("🏪 Outlet Performance Table").style(ST)

        sf = s.copy()
        if "outlet_name" in sf.columns:
            sf["outlet_name"] = sf["outlet_name"].fillna("").astype(str).str.strip()
            sf = sf[sf["outlet_name"] != ""]
        cs = sf[sf["periode"].astype(str) == str(cpv)].copy()
        cm = {}
        if not cs.empty and "outlet_name" in cs.columns:
            cs["_k"] = cs["outlet_name"].str.strip().str.lower()
            cm = (
                cs.drop_duplicates("_k", keep="last")
                .set_index("_k")
                .to_dict(orient="index")
            )
        mt2 = (
            sf.drop_duplicates("outlet_name", keep="last")
            .set_index("outlet_name")
            .to_dict(orient="index")
        )
        cpm = {}
        if cmv:
            cd = sf[sf["periode"].astype(str) == str(cmv)].copy()
            if not cd.empty:
                cd["_k"] = cd["outlet_name"].str.strip().str.lower()
                cpm = cd.set_index("_k").to_dict(orient="index")
        all_o = (
            sorted(sf["outlet_name"].dropna().astype(str).unique().tolist())
            if "outlet_name" in sf.columns
            else []
        )
        rows = []
        for name in all_o:
            k = name.strip().lower()
            act = k in cm
            r = cm.get(k, mt2.get(name, {}))
            if not act:
                continue
            st = str(r.get("outlet_status", ""))
            if st in ("Tidak Aktif", "", None):
                continue
            oms = float(r.get("total_revenue", 0) or 0)
            fot = int(r.get("foto_qty", 0) or 0)
            unl = int(r.get("unlock_qty", 0) or 0)
            prn = int(r.get("print_qty", 0) or 0)
            cnv = float(r.get("paid_per_photo_rate", 0) or 0)
            sta = str(r.get("outlet_status", ""))
            are = str(r.get("area", ""))
            has_cmp = cmv and k in cpm
            po = float(cpm[k].get("total_revenue", 0) or 0) if has_cmp else 0

            def _d(v, fn):
                if has_cmp:
                    df = float(r.get(v, 0) or 0) - float(cpm[k].get(v, 0) or 0)
                    if df > 0:
                        return f'<span class="tbl-gr">▲ {fn(df)}</span>'
                    elif df < 0:
                        return f'<span class="tbl-rd">▼ {fn(abs(df))}</span>'
                    else:
                        return '<span class="tbl-gd">● 0</span>'
                return '<span style="color:#585b70">—</span>'

            def _dp(v):
                if has_cmp:
                    cv = float(r.get(v, 0) or 0)
                    pv = float(cpm[k].get(v, 0) or 0)
                    d = cv - pv
                    if d > 0:
                        return f'<span class="tbl-gr">▲ +{_p(d)}</span>'
                    elif d < 0:
                        return f'<span class="tbl-rd">▼ {_p(d)}</span>'
                    else:
                        return '<span class="tbl-gd">● 0</span>'
                return '<span style="color:#585b70">—</span>'

            delta_oms = ""
            if has_cmp:
                if oms > po:
                    delta_oms = f'<span class="tbl-gr">▲ {_c(oms-po)}</span>'
                elif oms < po:
                    delta_oms = f'<span class="tbl-rd">▼ {_c(po-oms)}</span>'
                else:
                    delta_oms = f'<span class="tbl-gd">● {_c(oms)}</span>'
            rows.append(
                {
                    "Outlet": name,
                    "Area": are,
                    "Status": f'<span style="color:{SC.get(sta,"#94a3b8")};font-weight:600;">{sta}</span>',
                    "Omset": _c(oms),
                    "Delta Omset": delta_oms,
                    "Foto": _n(fot),
                    "Delta Foto": _d("foto_qty", lambda x: _n(int(x))),
                    "Unlock": _n(unl),
                    "Delta Unlock": _d("unlock_qty", lambda x: _n(int(x))),
                    "Print": _n(prn),
                    "Delta Print": _d("print_qty", lambda x: _n(int(x))),
                    "Conv": _p(cnv),
                    "Delta Conv": _dp("paid_per_photo_rate"),
                }
            )
        if rows:
            html_col_indices = []
            col_defs = []
            col_defs.append(
                {
                    "headerName": "Outlet",
                    "field": "Outlet",
                    "pinned": "left",
                    "minWidth": 200,
                    "flex": 3,
                    "sortable": True,
                    "filter": "agTextColumnFilter",
                    "floatingFilter": True,
                }
            )
            col_defs.append(
                {
                    "headerName": "Area",
                    "field": "Area",
                    "minWidth": 130,
                    "flex": 1,
                    "sortable": True,
                    "filter": "agTextColumnFilter",
                    "floatingFilter": True,
                }
            )
            col_defs.append(
                {
                    "headerName": "Status",
                    "field": "Status",
                    "minWidth": 110,
                    "flex": 1,
                    "sortable": True,
                    "filter": "agTextColumnFilter",
                    "floatingFilter": True,
                }
            )
            html_col_indices.append(len(col_defs) - 1)
            col_defs.append(
                {
                    "headerName": "Omset",
                    "field": "Omset",
                    "minWidth": 150,
                    "flex": 1,
                    "sortable": True,
                    "filter": "agNumberColumnFilter",
                    "floatingFilter": True,
                    "cellStyle": {"fontWeight": "700", "fontSize": "13px"},
                    ":comparator": "(a,b)=>{const x=a.split(' ').slice(-1)[0];const y=b.split(' ').slice(-1)[0];return Number(x?x.split('.').join(''):0)-Number(y?y.split('.').join(''):0)}",
                }
            )
            if cmv:
                col_defs.append(
                    {
                        "headerName": "Δ Omset",
                        "field": "Delta Omset",
                        "minWidth": 120,
                        "flex": 1,
                    }
                )
            col_defs.append(
                {
                    "headerName": "Foto",
                    "field": "Foto",
                    "minWidth": 90,
                    "flex": 1,
                    "sortable": True,
                    "filter": "agNumberColumnFilter",
                    "floatingFilter": True,
                    ":comparator": "(a,b)=>{const na=Number(a?a.split('.').join(''):0);const nb=Number(b?b.split('.').join(''):0);return na-nb}",
                }
            )
            if cmv:
                col_defs.append(
                    {
                        "headerName": "Δ Foto",
                        "field": "Delta Foto",
                        "minWidth": 100,
                        "flex": 1,
                    }
                )
            col_defs.append(
                {
                    "headerName": "Unlock",
                    "field": "Unlock",
                    "minWidth": 90,
                    "flex": 1,
                    "sortable": True,
                    "filter": "agNumberColumnFilter",
                    "floatingFilter": True,
                    ":comparator": "(a,b)=>{const na=Number(a?a.split('.').join(''):0);const nb=Number(b?b.split('.').join(''):0);return na-nb}",
                }
            )
            if cmv:
                col_defs.append(
                    {
                        "headerName": "Δ Unlock",
                        "field": "Delta Unlock",
                        "minWidth": 100,
                        "flex": 1,
                    }
                )
            col_defs.append(
                {
                    "headerName": "Print",
                    "field": "Print",
                    "minWidth": 90,
                    "flex": 1,
                    "sortable": True,
                    "filter": "agNumberColumnFilter",
                    "floatingFilter": True,
                    ":comparator": "(a,b)=>{const na=Number(a?a.split('.').join(''):0);const nb=Number(b?b.split('.').join(''):0);return na-nb}",
                }
            )
            if cmv:
                col_defs.append(
                    {
                        "headerName": "Δ Print",
                        "field": "Delta Print",
                        "minWidth": 100,
                        "flex": 1,
                    }
                )
            col_defs.append(
                {
                    "headerName": "Conv",
                    "field": "Conv",
                    "minWidth": 100,
                    "flex": 1,
                    "sortable": True,
                    "filter": "agNumberColumnFilter",
                    "floatingFilter": True,
                    ":comparator": "(a,b)=>{const na=Number(a?a.replace('%','').replace(',','.'):0);const nb=Number(b?b.replace('%','').replace(',','.'):0);return na-nb}",
                }
            )
            if cmv:
                col_defs.append(
                    {
                        "headerName": "Δ Conv",
                        "field": "Delta Conv",
                        "minWidth": 100,
                        "flex": 1,
                    }
                )
            ui.aggrid(
                {
                    "columnDefs": col_defs,
                    "rowData": rows,
                    "pagination": True,
                    "paginationPageSize": 50,
                    "paginationPageSizeSelector": [25, 50, 100],
                    "domLayout": "autoHeight",
                    "defaultColDef": {"resizable": True},
                    "animateRows": True,
                    "rowHeight": 40,
                    "headerHeight": 40,
                    "enableCellTextSelection": True,
                },
                theme="balham",
                html_columns=html_col_indices,
            ).classes("w-full ag-theme-balham-dark").style(
                "height: auto; min-height: 300px;"
            )
        else:
            ui.label("Tidak ada outlet.").classes("text-gray-400 italic")

        if md.empty:
            return
        with ui.row().classes("w-full gap-4 mb-6"):
            with ui.card().classes("flex-1").style(CARD):
                ui.label("📊 Distribusi Status Outlet").style(ST)
                if "outlet_status" in md.columns:
                    sc = md["outlet_status"].value_counts()
                    ui.echart(
                        {
                            "tooltip": {
                                "trigger": "item",
                                "formatter": "{b}:{c}({d}%)",
                            },
                            "color": [SC.get(s, "#94a3b8") for s in sc.index.tolist()],
                            "series": [
                                {
                                    "type": "pie",
                                    "radius": ["40%", "70%"],
                                    "center": ["50%", "50%"],
                                    "data": [
                                        {"name": k, "value": v}
                                        for k, v in zip(
                                            sc.index.tolist(), sc.values.tolist()
                                        )
                                    ],
                                    "label": {
                                        "color": "#cdd6f4",
                                        "fontSize": 11,
                                        "formatter": "{b}:{c}",
                                    },
                                    "itemStyle": {
                                        "borderColor": "#1e1e2e",
                                        "borderWidth": 2,
                                    },
                                }
                            ],
                            **{
                                k: v
                                for k, v in ECHART.items()
                                if k not in ("xAxis", "yAxis")
                            },
                        }
                    ).classes("w-full h-[300px]")
                else:
                    ui.label("Tidak ada data.").classes("text-gray-400 italic")
            with ui.card().classes("flex-1").style(CARD):
                tp = a.get_top_performers(md, 5)
                wr = a.get_worst_performers(md, 10)
                ui.label("🏆 Top 5").style(ST)
                if not tp.empty:
                    for _, r in tp.iterrows():
                        st = str(r.get("outlet_status", ""))
                        sc = SC.get(st, "#94a3b8")
                        with ui.row().classes(
                            "items-center w-full py-1 px-2 rounded-lg"
                        ).style("background:#181825"):
                            ui.label(r["outlet_name"]).classes(
                                "text-sm text-white flex-1"
                            )
                            ui.label(st).style(
                                f"color:{sc};font-weight:bold;font-size:0.8rem"
                            )
                            ui.label(a.format_currency(r["total_revenue"])).classes(
                                "text-sm text-green-400 ml-2"
                            )
                else:
                    ui.label("Tidak ada data.").classes("text-gray-400 italic text-xs")
                ui.separator().classes("my-3")
                ui.label("⬇️ 10 Terjelek").style(ST)
                if not wr.empty:
                    for _, r in wr.iterrows():
                        st = str(r.get("outlet_status", ""))
                        sc = SC.get(st, "#94a3b8")
                        with ui.row().classes(
                            "items-center w-full py-1 px-2 rounded-lg"
                        ).style("background:#181825;border-left:3px solid #ef4444;"):
                            ui.label(r["outlet_name"]).classes(
                                "text-sm text-white flex-1"
                            )
                            ui.label(st).style(
                                f"color:{sc};font-weight:bold;font-size:0.8rem"
                            )
                            ui.label(a.format_currency(r["total_revenue"])).classes(
                                "text-sm text-red-400 ml-2"
                            )
                else:
                    ui.label("Tidak ada data.").classes("text-gray-400 italic text-xs")
        with ui.row().classes("w-full gap-4 mb-6"):
            with ui.card().classes("flex-[2]").style(CARD):
                ui.label("💹 Revenue by Outlet").style(ST)
                if not md.empty and "total_revenue" in md.columns:
                    tp = md.nlargest(10, "total_revenue")
                    ui.echart(
                        {
                            "tooltip": {
                                "trigger": "axis",
                                "formatter": "{b}<br/>Revenue:Rp{c}",
                            },
                            "grid": {
                                "left": "3%",
                                "right": "4%",
                                "bottom": "15%",
                                "containLabel": True,
                            },
                            "xAxis": {
                                "type": "category",
                                "data": tp["outlet_name"].tolist(),
                                "axisLabel": {
                                    "rotate": 35,
                                    "fontSize": 10,
                                    "color": "#a6adc8",
                                },
                                "axisLine": {"lineStyle": {"color": "#45475a"}},
                            },
                            "yAxis": {
                                "type": "value",
                                "axisLabel": {
                                    "color": "#a6adc8",
                                    "formatter": "Rp{value}",
                                },
                                "splitLine": {"lineStyle": {"color": "#313244"}},
                            },
                            "series": [
                                {
                                    "name": "Revenue",
                                    "type": "bar",
                                    "data": tp["total_revenue"].tolist(),
                                    "itemStyle": {
                                        "color": {
                                            "type": "linear",
                                            "x": 0,
                                            "y": 0,
                                            "x2": 0,
                                            "y2": 1,
                                            "colorStops": [
                                                {"offset": 0, "color": "#89b4fa"},
                                                {"offset": 1, "color": "#45475a"},
                                            ],
                                        }
                                    },
                                    "barMaxWidth": 40,
                                    "label": {
                                        "show": True,
                                        "position": "top",
                                        "formatter": "{@value}",
                                        "color": "#cdd6f4",
                                        "fontSize": 10,
                                    },
                                }
                            ],
                            **ECHART,
                        }
                    ).classes("w-full h-[350px]")
                else:
                    ui.label("Tidak ada data.").classes("text-gray-400 italic")
            with ui.card().classes("flex-1").style(CARD):
                ui.label("🔄 Conversion Funnel").style(ST)
                ft = int(md["foto_qty"].sum()) if "foto_qty" in md.columns else 0
                ul = int(md["unlock_qty"].sum()) if "unlock_qty" in md.columns else 0
                pr = int(md["print_qty"].sum()) if "print_qty" in md.columns else 0
                ui.echart(
                    {
                        "tooltip": {"trigger": "item", "formatter": "{b}:{c}"},
                        "series": [
                            {
                                "type": "funnel",
                                "left": "10%",
                                "top": 20,
                                "bottom": 20,
                                "width": "80%",
                                "min": 0,
                                "max": ft,
                                "minSize": "0%",
                                "maxSize": "100%",
                                "sort": "descending",
                                "gap": 2,
                                "label": {
                                    "show": True,
                                    "position": "inside",
                                    "color": "#fff",
                                    "fontSize": 12,
                                    "formatter": "{b}:{c}",
                                },
                                "itemStyle": {
                                    "borderColor": "#1e1e2e",
                                    "borderWidth": 2,
                                },
                                "data": [
                                    {
                                        "value": ft,
                                        "name": "📸 Foto",
                                        "itemStyle": {"color": "#89b4fa"},
                                    },
                                    {
                                        "value": ul,
                                        "name": "🔓 Unlock",
                                        "itemStyle": {"color": "#f9e2af"},
                                    },
                                    {
                                        "value": pr,
                                        "name": "🖨️ Print",
                                        "itemStyle": {"color": "#a6e3a1"},
                                    },
                                ],
                            }
                        ],
                        **{
                            k: v
                            for k, v in ECHART.items()
                            if k not in ("xAxis", "yAxis")
                        },
                    }
                ).classes("w-full h-[300px]")
        ui.separator().classes("my-4")
        ui.label("💡 Key Insights").style(ST)
        ins = a.get_insights(md)
        if ins:
            for i in ins:
                with ui.card().classes("w-full mb-2").style(
                    "background:#181825;border-left:3px solid #89b4fa;border-radius:8px;padding:12px 16px;"
                ):
                    ui.label(i).classes("text-sm text-gray-300")
        else:
            ui.label("Tidak ada insight.").classes("text-gray-400 italic")
        ui.separator().classes("my-6")
        ols = (
            md["outlet_name"].dropna().unique().tolist()
            if "outlet_name" in md.columns
            else []
        )
        if ols:
            tr = a.build_trend_table(_ff, _cp, ols, 12)
            if tr["has_data"]:
                vc = tr["value_cols"]
                ad = tr["active_df"]
                idf = tr["inactive_df"]
                with ui.expansion(
                    "📆 Tren Omset Outlet (12 Bulan)", icon="trending_up"
                ).classes("w-full mb-4"):
                    if not ad.empty:
                        ui.label("Outlet Aktif").classes(
                            "text-sm font-semibold text-green-400 mt-2 mb-2"
                        )
                        rr = []
                        for _, rd in ad.iterrows():
                            d = {
                                "Outlet": str(rd.get("Outlet", "")),
                                "Rata-rata": str(rd.get("Rata-rata", "Rp 0")),
                            }
                            for p in vc:
                                d[p] = str(rd.get(p, "Rp 0"))
                            rr.append(d)
                        tr_cols = [
                            {
                                "name": "Outlet",
                                "label": "Outlet",
                                "field": "Outlet",
                                "align": "left",
                                "sortable": True,
                            }
                        ]
                        tr_cols.append(
                            {
                                "name": "Rata-rata",
                                "label": "Rata-rata",
                                "field": "Rata-rata",
                                "align": "right",
                                "sortable": True,
                            }
                        )
                        for p in vc:
                            tr_cols.append(
                                {"name": p, "label": p, "field": p, "align": "right"}
                            )
                        # Color each month value green/red vs prev month
                        trend_html_indices = list(range(2, 2 + len(vc)))
                        for ri in rr:
                            months = list(vc)
                            prev_val = None
                            for i, p in enumerate(months):
                                raw = ri.get(p, "Rp 0")
                                try:
                                    s = raw.replace("Rp ", "").strip()
                                    cur = float(s.replace(".", "").replace(",", ""))
                                except:
                                    cur = 0
                                # For latest month, compare with the next month
                                cmp_val = prev_val
                                if i == 0 and len(months) > 1:
                                    try:
                                        s2 = (
                                            ri.get(months[1], "Rp 0")
                                            .replace("Rp ", "")
                                            .strip()
                                        )
                                        cmp_val = float(
                                            s2.replace(".", "").replace(",", "")
                                        )
                                    except:
                                        pass
                                if cmp_val is not None and cur != 0:
                                    if cur > cmp_val:
                                        ri[p] = (
                                            '<span style="color:#a6e3a1;font-weight:600">'
                                            + raw
                                            + "</span>"
                                        )
                                    elif cur < cmp_val:
                                        ri[p] = (
                                            '<span style="color:#f38ba8;font-weight:600">'
                                            + raw
                                            + "</span>"
                                        )
                                prev_val = cur
                        ui.aggrid(
                            {
                                "columnDefs": [
                                    {
                                        "headerName": "Outlet",
                                        "field": "Outlet",
                                        "pinned": "left",
                                        "minWidth": 180,
                                        "flex": 2,
                                        "sortable": True,
                                    },
                                    {
                                        "headerName": "Rata-rata",
                                        "field": "Rata-rata",
                                        "minWidth": 130,
                                        "flex": 1,
                                        "sortable": True,
                                    },
                                ]
                                + [
                                    {
                                        "headerName": p,
                                        "field": p,
                                        "minWidth": 110,
                                        "flex": 1,
                                    }
                                    for p in vc
                                ],
                                "rowData": rr,
                                "pagination": True,
                                "paginationPageSize": 10,
                                "domLayout": "autoHeight",
                                "defaultColDef": {"resizable": True},
                                "animateRows": True,
                                "rowHeight": 40,
                                "headerHeight": 40,
                            },
                            theme="balham",
                            html_columns=trend_html_indices,
                        ).classes("w-full ag-theme-balham-dark").style(
                            "height: auto; min-height: 200px;"
                        )
                    else:
                        ui.label("Tidak ada outlet aktif.").classes(
                            "text-gray-400 italic"
                        )
                    if not idf.empty:
                        ui.label("Outlet Tidak Aktif").classes(
                            "text-sm font-semibold text-gray-400 mt-2 mb-2"
                        )
                        rr = []
                        for _, rd in idf.iterrows():
                            d = {
                                "Outlet": str(rd.get("Outlet", "")),
                                "Rata-rata": str(rd.get("Rata-rata", "Rp 0")),
                            }
                            for p in vc:
                                d[p] = str(rd.get(p, "Rp 0"))
                            rr.append(d)
                        trend_html_indices2 = list(range(2, 2 + len(vc)))
                        for ri in rr:
                            months = list(vc)
                            prev_val = None
                            for i, p in enumerate(months):
                                raw = ri.get(p, "Rp 0")
                                try:
                                    s = raw.replace("Rp ", "").strip()
                                    cur = float(s.replace(".", "").replace(",", ""))
                                except:
                                    cur = 0
                                cmp_val = prev_val
                                if i == 0 and len(months) > 1:
                                    try:
                                        s2 = (
                                            ri.get(months[1], "Rp 0")
                                            .replace("Rp ", "")
                                            .strip()
                                        )
                                        cmp_val = float(
                                            s2.replace(".", "").replace(",", "")
                                        )
                                    except:
                                        pass
                                if cmp_val is not None and cur != 0:
                                    if cur > cmp_val:
                                        ri[p] = (
                                            '<span style="color:#a6e3a1;font-weight:600">'
                                            + raw
                                            + "</span>"
                                        )
                                    elif cur < cmp_val:
                                        ri[p] = (
                                            '<span style="color:#f38ba8;font-weight:600">'
                                            + raw
                                            + "</span>"
                                        )
                                prev_val = cur
                        ui.aggrid(
                            {
                                "columnDefs": [
                                    {
                                        "headerName": "Outlet",
                                        "field": "Outlet",
                                        "pinned": "left",
                                        "minWidth": 180,
                                        "flex": 2,
                                        "sortable": True,
                                    },
                                    {
                                        "headerName": "Rata-rata",
                                        "field": "Rata-rata",
                                        "minWidth": 130,
                                        "flex": 1,
                                        "sortable": True,
                                    },
                                ]
                                + [
                                    {
                                        "headerName": p,
                                        "field": p,
                                        "minWidth": 110,
                                        "flex": 1,
                                    }
                                    for p in vc
                                ],
                                "rowData": rr,
                                "pagination": True,
                                "paginationPageSize": 5,
                                "domLayout": "autoHeight",
                                "defaultColDef": {"resizable": True},
                                "animateRows": True,
                                "rowHeight": 40,
                                "headerHeight": 40,
                            },
                            theme="balham",
                            html_columns=trend_html_indices2,
                        ).classes("w-full ag-theme-balham-dark").style(
                            "height: auto; min-height: 200px;"
                        )
                    ui.label(
                        "Nilai kosong = 0. Rata-rata dari 12 bulan, hanya omset > 0 dihitung."
                    ).classes("text-[10px] text-gray-500 italic")
        else:
            ui.label("Tidak ada outlet.").classes("text-gray-400 italic")


def _get_last_update_info() -> dict:
    """Get data freshness info from cache file timestamps."""
    from pathlib import Path
    
    info = {"cache_time": None, "latest_date": None, "hours_old": None, "is_fresh": True}
    
    # Dashboard summary freshness
    cache_path = Path(__file__).resolve().parent.parent.parent
    cache_path = cache_path / "streamlit_template" / "data" / "api_cache" / "dashboard_summary.json"
    if cache_path.exists():
        mtime = os.path.getmtime(str(cache_path))
        dt = datetime.fromtimestamp(mtime)
        info["cache_time"] = dt.strftime("%d %b %Y %H:%M")
        info["hours_old"] = (time.time() - mtime) / 3600
        info["is_fresh"] = info["hours_old"] <= 8
    
    # Latest date in raw data
    raw_dir = cache_path.parent / "raw_by_month"
    if raw_dir.exists():
        files = sorted(raw_dir.glob("*.json"))
        if files:
            try:
                import json as _j
                with open(str(files[-1])) as _f:
                    _data = _j.load(_f)
                if _data:
                    _dates = sorted(set(t.get("date","") for t in _data if t.get("date")))
                    if _dates:
                        info["latest_date"] = _dates[-1]
            except Exception:
                pass
    
    return info


def set_filters(df, fdf, cp, cmp, area="Semua", kat="Semua", tip="Semua"):
    global _df, _ff, _cp, _cmp
    _df = df
    _ff = fdf
    _cp = cp
    _cmp = cmp
    if _act_sel is not None:
        periods = _periods
        _act_sel.options = periods
        _act_sel.value = cp or periods[-1]
        cmp_opts = ["-"] + [p for p in periods if p != (cp or periods[-1])]
        _cmp_sel.options = cmp_opts
        _cmp_sel.value = cmp if cmp in cmp_opts else "-"
    if _content is not None:
        _build_outlet_table(_cp, _cmp)


def create_page(c):
    global _content, _act_sel, _cmp_sel, _periods, _a
    _a = get_adapter()
    _content = ui.column().classes("w-full")
    with c:
        ui.add_head_html(TBL_CSS)
        with ui.row().classes("w-full items-center gap-3 mb-2"):
            ui.label("📸").classes("text-3xl")
            ui.label("difotoin.id").classes("text-2xl font-bold text-white")
            ui.label("— Dashboard").classes("text-lg text-gray-400")
            ui.label(f"• {datetime.now().strftime('%d %b %Y %H:%M')}").classes(
                "text-xs text-gray-500 ml-auto"
            )
        ui.separator().classes("mb-3")

        # ── Last Update Info ──
        _info = _get_last_update_info()
        _info_row = ui.row().classes("w-full items-center gap-2 mb-3")
        with _info_row:
            cache_str = _info.get("cache_time") or "-"
            date_str = _info.get("latest_date") or "-"
            hours = _info.get("hours_old")
            if hours is not None and hours > 24:
                color = "#ef4444"  # red — stale
                icon = "🔴"
            elif hours is not None and hours > 8:
                color = "#f59e0b"  # amber — aging
                icon = "🟡"
            else:
                color = "#22c55e"  # green — fresh
                icon = "🟢"
            ui.label(f"{icon} Update cache: {cache_str}").classes("text-xs").style(f"color: {color}")
            ui.label(f"• Data terakhir: {date_str}").classes("text-xs text-gray-400")
            if hours is not None:
                ui.label(f"({hours:.0f}h yang lalu)").classes("text-xs").style(f"color: {color}")
            else:
                ui.label("(belum ada sync)").classes("text-xs text-red-400")
            ui.label("• Sync tiap 1 jam").classes("text-xs text-gray-500 ml-auto")
        ui.separator().classes("mb-6")

        # STABLE dropdowns (never destroyed)
        data = _ff if _ff is not None else _a.load_full_data()
        if data is not None and not data.empty:
            _periods = (
                sorted(data["periode"].dropna().astype(str).unique().tolist())
                if "periode" in data.columns
                else []
            )
        if not _periods:
            _periods = ["-"]
        with ui.row().classes("w-full items-center gap-4 mb-3 flex-wrap"):
            ui.label("Periode:").classes("text-xs text-gray-400")
            _act_sel = (
                ui.select(_periods, value=_cp or _periods[-1], label="Bulan Aktif")
                .props("dense outlined dark")
                .classes("w-40")
            )
            ui.label("Bandingkan:").classes("text-xs text-gray-400")
            _cmp_sel = (
                ui.select(["-"], value="-", label="Bandingkan")
                .props("dense outlined dark")
                .classes("w-40")
            )

        # Init compare options
        if _act_sel.value:
            _cmp_sel.options = ["-"] + [p for p in _periods if p != _act_sel.value]
            _cmp_sel.value = _cmp if _cmp in _cmp_sel.options else "-"

        # Wire callbacks
        def _on_change():
            global _cp, _cmp
            _cp = _act_sel.value
            _cmp = None if _cmp_sel.value == "-" else _cmp_sel.value
            _build_outlet_table(_cp, _cmp)

        _act_sel.on("update:model-value", _on_change)
        _cmp_sel.on("update:model-value", _on_change)

        # Build initial content
        _build_outlet_table(_cp or (_periods[-1] if _periods else None), _cmp)



_EXPORT_HTML_RE = re.compile(r"<[^>]+>")


def _clean_export_value(value):
    """Strip AG Grid HTML coloring so Excel/PDF berisi angka/text bersih."""
    if value is None:
        return ""
    text = unescape(_EXPORT_HTML_RE.sub("", str(value)))
    return " ".join(text.split())


def _trend_export_columns(value_cols):
    return ["Outlet", "Rata-rata"] + list(value_cols)


def _trend_export_rows(rows, value_cols):
    columns = _trend_export_columns(value_cols)
    clean_rows = []
    for row in rows or []:
        clean_rows.append([_clean_export_value(row.get(col, "")) for col in columns])
    return columns, clean_rows


def _export_tmp_path(prefix, ext):
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return os.path.join(tempfile.gettempdir(), f"{prefix}_{stamp}.{ext}")


def _download_trend_excel(rows, value_cols, label):
    if not rows:
        ui.notify("Pilih outlet dulu yang mau didownload.", type="warning")
        return
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill

        columns, clean_rows = _trend_export_rows(rows, value_cols)
        wb = Workbook()
        ws = wb.active
        ws.title = "Tren Omset 12 Bulan"
        ws.append(["Tren Omset Outlet (12 Bulan)"])
        ws.append(["Kategori", label])
        ws.append(["Generated", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
        ws.append([])
        ws.append(columns)
        for row in clean_rows:
            ws.append(row)

        header_row = 5
        for cell in ws[header_row]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1E1E2E")
            cell.alignment = Alignment(horizontal="center")
        for row in ws.iter_rows(min_row=header_row + 1):
            for cell in row:
                cell.alignment = Alignment(horizontal="right" if cell.column > 1 else "left")
        ws.freeze_panes = "C6"
        for idx, col in enumerate(columns, start=1):
            letter = ws.cell(row=header_row, column=idx).column_letter
            if col == "Outlet":
                ws.column_dimensions[letter].width = 32
            elif col == "Rata-rata":
                ws.column_dimensions[letter].width = 16
            else:
                ws.column_dimensions[letter].width = 14

        path = _export_tmp_path("tren_omset_12_bulan", "xlsx")
        wb.save(path)
        ui.download(path)
        ui.notify(f"Download Excel siap: {len(clean_rows)} outlet", type="positive")
    except Exception as exc:
        ui.notify(f"Gagal export Excel: {exc}", type="negative")


def _pdf_escape(text):
    return str(text).replace("\\", "\\\\").replace("(", "\\(").replace(")", "\\)")


def _pdf_text_cmd(x, y, text, size=7):
    return f"BT /F1 {size} Tf 1 0 0 1 {x:.1f} {y:.1f} Tm ({_pdf_escape(text)}) Tj ET"


def _pdf_line_cmd(x1, y1, x2, y2):
    return f"0.35 w {x1:.1f} {y1:.1f} m {x2:.1f} {y2:.1f} l S"


def _write_simple_pdf(path, title, columns, rows):
    # A4 landscape; compact fixed-width report tanpa dependency eksternal.
    width, height = 842, 595
    margin = 24
    col_widths = [155, 82]
    rest = max(42, (width - margin * 2 - sum(col_widths)) / max(1, len(columns) - 2))
    col_widths += [rest] * (len(columns) - 2)

    def clip(text, col_w):
        text = _clean_export_value(text)
        max_chars = max(4, int(col_w / 3.7))
        return text if len(text) <= max_chars else text[: max_chars - 1] + "…"

    pages = []
    row_h = 13
    data_start_y = height - 82
    min_y = margin + 20
    rows_per_page = max(1, int((data_start_y - min_y) / row_h))
    chunks = [rows[i : i + rows_per_page] for i in range(0, len(rows), rows_per_page)] or [[]]
    total_pages = len(chunks)

    for page_no, chunk in enumerate(chunks, start=1):
        cmds = []
        cmds.append(_pdf_text_cmd(margin, height - 32, title, 11))
        cmds.append(_pdf_text_cmd(margin, height - 48, f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | Rows: {len(rows)}", 7))
        cmds.append(_pdf_text_cmd(width - 88, height - 48, f"Page {page_no}/{total_pages}", 7))
        y = height - 68
        x = margin
        for col, cw in zip(columns, col_widths):
            cmds.append(_pdf_text_cmd(x + 2, y, clip(col, cw), 6.5))
            x += cw
        cmds.append(_pdf_line_cmd(margin, y - 4, width - margin, y - 4))
        y -= row_h
        for row in chunk:
            x = margin
            for val, cw in zip(row, col_widths):
                cmds.append(_pdf_text_cmd(x + 2, y, clip(val, cw), 6.2))
                x += cw
            cmds.append(_pdf_line_cmd(margin, y - 4, width - margin, y - 4))
            y -= row_h
        pages.append("\n".join(cmds))

    objects = ["<< /Type /Catalog /Pages 2 0 R >>"]
    kids = []
    for i, commands in enumerate(pages):
        page_obj = 3 + i * 2
        content_obj = page_obj + 1
        kids.append(f"{page_obj} 0 R")
        stream = commands.encode("latin-1", "replace")
        objects.append(f"<< /Type /Page /Parent 2 0 R /MediaBox [0 0 {width} {height}] /Resources << /Font << /F1 {3 + len(pages) * 2} 0 R >> >> /Contents {content_obj} 0 R >>")
        objects.append(f"<< /Length {len(stream)} >>\nstream\n{commands}\nendstream")
    objects.insert(1, f"<< /Type /Pages /Kids [{' '.join(kids)}] /Count {len(kids)} >>")
    objects.append("<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica >>")

    out = ["%PDF-1.4\n"]
    offsets = [0]
    for idx, obj in enumerate(objects, start=1):
        offsets.append(sum(len(part.encode("latin-1", "replace")) for part in out))
        out.append(f"{idx} 0 obj\n{obj}\nendobj\n")
    xref_at = sum(len(part.encode("latin-1", "replace")) for part in out)
    out.append(f"xref\n0 {len(objects)+1}\n0000000000 65535 f \n")
    for off in offsets[1:]:
        out.append(f"{off:010d} 00000 n \n")
    out.append(f"trailer\n<< /Size {len(objects)+1} /Root 1 0 R >>\nstartxref\n{xref_at}\n%%EOF\n")
    with open(path, "wb") as fh:
        fh.write("".join(out).encode("latin-1", "replace"))


def _download_trend_pdf(rows, value_cols, label):
    if not rows:
        ui.notify("Pilih outlet dulu yang mau didownload.", type="warning")
        return
    try:
        columns, clean_rows = _trend_export_rows(rows, value_cols)
        path = _export_tmp_path("tren_omset_12_bulan", "pdf")
        _write_simple_pdf(path, f"Tren Omset Outlet (12 Bulan) - {label}", columns, clean_rows)
        ui.download(path)
        ui.notify(f"Download PDF siap: {len(clean_rows)} outlet", type="positive")
    except Exception as exc:
        ui.notify(f"Gagal export PDF: {exc}", type="negative")


def _build_outlet_table(cpv, cmv):
    """Rebuild just the outlet table content."""
    global _a
    a = _a
    src = _ff
    if _content is None:
        return
    _content.clear()
    if src is None or src.empty:
        return
    s = src.copy()
    for cl in [
        "total_revenue",
        "foto_qty",
        "unlock_qty",
        "print_qty",
        "paid_per_photo_rate",
    ]:
        if cl in s.columns:
            s[cl] = pd.to_numeric(s[cl], errors="coerce").fillna(0)
    if _cp and "periode" in s.columns:
        md = s[s["periode"].astype(str) == str(_cp)].copy()
    else:
        md = s.copy()

    with _content:
        mt = a.calculate_metrics(md)
        with ui.row().classes("w-full gap-4 mb-6"):
            for lb, vl, cl in [
                ("💰 Revenue", a.format_currency(mt["total_revenue"]), "#89b4fa"),
                ("🏪 Outlets", _n(mt["total_outlets"]), "#a6e3a1"),
                ("📈 Avg Conv Rate", f"{mt['avg_conversion']:.1f}%", "#f9e2af"),
                ("📸 Photos", _n(mt["total_photos"]), "#f38ba8"),
            ]:
                with ui.card().classes("flex-1 min-w-[150px]").style(CARD):
                    ui.label(lb).style(ML)
                    ui.label(vl).style(MV)
                    ui.label(f"Periode: {_cp or 'Semua'}").classes(
                        "text-[10px] text-gray-600 mt-1"
                    )
        ui.separator().classes("mb-4")
        ui.label("🏪 Outlet Performance Table").style(ST)

        sf = s.copy()
        if "outlet_name" in sf.columns:
            sf["outlet_name"] = sf["outlet_name"].fillna("").astype(str).str.strip()
            sf = sf[sf["outlet_name"] != ""]
        cs = sf[sf["periode"].astype(str) == str(cpv)].copy()
        cm = {}
        if not cs.empty and "outlet_name" in cs.columns:
            cs["_k"] = cs["outlet_name"].str.strip().str.lower()
            cm = (
                cs.drop_duplicates("_k", keep="last")
                .set_index("_k")
                .to_dict(orient="index")
            )
        mt2 = (
            sf.drop_duplicates("outlet_name", keep="last")
            .set_index("outlet_name")
            .to_dict(orient="index")
        )
        cpm = {}
        if cmv:
            cd = sf[sf["periode"].astype(str) == str(cmv)].copy()
            if not cd.empty:
                cd["_k"] = cd["outlet_name"].str.strip().str.lower()
                cpm = cd.set_index("_k").to_dict(orient="index")
        all_o = (
            sorted(sf["outlet_name"].dropna().astype(str).unique().tolist())
            if "outlet_name" in sf.columns
            else []
        )
        rows = []
        for name in all_o:
            k = name.strip().lower()
            act = k in cm
            r = cm.get(k, mt2.get(name, {}))
            if not act:
                continue
            st = str(r.get("outlet_status", ""))
            if st in ("Tidak Aktif", "", None):
                continue
            oms = float(r.get("total_revenue", 0) or 0)
            fot = int(r.get("foto_qty", 0) or 0)
            unl = int(r.get("unlock_qty", 0) or 0)
            prn = int(r.get("print_qty", 0) or 0)
            cnv = float(r.get("paid_per_photo_rate", 0) or 0)
            are = str(r.get("area", ""))
            has_cmp = cmv and k in cpm
            po = float(cpm[k].get("total_revenue", 0) or 0) if has_cmp else 0

            def _d(v, fn):
                if has_cmp:
                    df = float(r.get(v, 0) or 0) - float(cpm[k].get(v, 0) or 0)
                    if df > 0:
                        return f'<span class="tbl-gr">▲ {fn(df)}</span>'
                    elif df < 0:
                        return f'<span class="tbl-rd">▼ {fn(abs(df))}</span>'
                    else:
                        return '<span class="tbl-gd">● 0</span>'
                return '<span style="color:#585b70">—</span>'

            def _dp(v):
                if has_cmp:
                    cv = float(r.get(v, 0) or 0)
                    pv = float(cpm[k].get(v, 0) or 0)
                    d = cv - pv
                    if d > 0:
                        return f'<span class="tbl-gr">▲ +{_p(d)}</span>'
                    elif d < 0:
                        return f'<span class="tbl-rd">▼ {_p(d)}</span>'
                    else:
                        return '<span class="tbl-gd">● 0</span>'
                return '<span style="color:#585b70">—</span>'

            delta_oms = ""
            if has_cmp:
                if oms > po:
                    delta_oms = f'<span class="tbl-gr">▲ {_c(oms-po)}</span>'
                elif oms < po:
                    delta_oms = f'<span class="tbl-rd">▼ {_c(po-oms)}</span>'
                else:
                    delta_oms = f'<span class="tbl-gd">● {_c(oms)}</span>'
            rows.append(
                {
                    "Outlet": name,
                    "Area": are,
                    "Omset": _c(oms),
                    "Delta Omset": delta_oms,
                    "Foto": _n(fot),
                    "Delta Foto": _d("foto_qty", lambda x: _n(int(x))),
                    "Unlock": _n(unl),
                    "Delta Unlock": _d("unlock_qty", lambda x: _n(int(x))),
                    "Print": _n(prn),
                    "Delta Print": _d("print_qty", lambda x: _n(int(x))),
                    "Conv": _p(cnv),
                    "Delta Conv": _dp("paid_per_photo_rate"),
                }
            )
        if rows:
            html_col_indices = []
            col_defs = []
            col_defs.append(
                {
                    "headerName": "Outlet",
                    "field": "Outlet",
                    "pinned": "left",
                    "minWidth": 200,
                    "flex": 3,
                    "sortable": True,
                    "filter": "agTextColumnFilter",
                    "floatingFilter": True,
                }
            )
            col_defs.append(
                {
                    "headerName": "Area",
                    "field": "Area",
                    "minWidth": 130,
                    "flex": 1,
                    "sortable": True,
                    "filter": "agTextColumnFilter",
                    "floatingFilter": True,
                }
            )
            col_defs.append(
                {
                    "headerName": "Omset",
                    "field": "Omset",
                    "minWidth": 150,
                    "flex": 1,
                    "sortable": True,
                    "filter": "agNumberColumnFilter",
                    "floatingFilter": True,
                    "cellStyle": {"fontWeight": "700", "fontSize": "13px"},
                    ":comparator": "(a,b)=>{const x=a.split(' ').slice(-1)[0];const y=b.split(' ').slice(-1)[0];return Number(x?x.split('.').join(''):0)-Number(y?y.split('.').join(''):0)}",
                }
            )
            if cmv:
                col_defs.append(
                    {
                        "headerName": "Δ Omset",
                        "field": "Delta Omset",
                        "minWidth": 120,
                        "flex": 1,
                        "sortable": True,
                        ":comparator": "(a,b)=>{const p=v=>{if(v==null)return 0;const s=String(v).replace(/<[^>]*>/g,'');const n=parseFloat(s.replace(/[^0-9.,]/g,'').replace(/\\./g,'').replace(',','.'));return (s.indexOf('▼')>=0?-1:1)*(isNaN(n)?0:n)};return p(a)-p(b)}",
                    }
                )
                html_col_indices.append(len(col_defs) - 1)
            col_defs.append(
                {
                    "headerName": "Foto",
                    "field": "Foto",
                    "minWidth": 90,
                    "flex": 1,
                    "sortable": True,
                    "filter": "agNumberColumnFilter",
                    "floatingFilter": True,
                    ":comparator": "(a,b)=>{const na=Number(a?a.split('.').join(''):0);const nb=Number(b?b.split('.').join(''):0);return na-nb}",
                }
            )
            if cmv:
                col_defs.append(
                    {
                        "headerName": "Δ Foto",
                        "field": "Delta Foto",
                        "minWidth": 100,
                        "flex": 1,
                        "sortable": True,
                        ":comparator": "(a,b)=>{const p=v=>{if(v==null)return 0;const s=String(v).replace(/<[^>]*>/g,'');const n=parseFloat(s.replace(/[^0-9.,]/g,'').replace(/\\./g,'').replace(',','.'));return (s.indexOf('▼')>=0?-1:1)*(isNaN(n)?0:n)};return p(a)-p(b)}",
                    }
                )
                html_col_indices.append(len(col_defs) - 1)
            col_defs.append(
                {
                    "headerName": "Unlock",
                    "field": "Unlock",
                    "minWidth": 90,
                    "flex": 1,
                    "sortable": True,
                    "filter": "agNumberColumnFilter",
                    "floatingFilter": True,
                    ":comparator": "(a,b)=>{const na=Number(a?a.split('.').join(''):0);const nb=Number(b?b.split('.').join(''):0);return na-nb}",
                }
            )
            if cmv:
                col_defs.append(
                    {
                        "headerName": "Δ Unlock",
                        "field": "Delta Unlock",
                        "minWidth": 100,
                        "flex": 1,
                        "sortable": True,
                        ":comparator": "(a,b)=>{const p=v=>{if(v==null)return 0;const s=String(v).replace(/<[^>]*>/g,'');const n=parseFloat(s.replace(/[^0-9.,]/g,'').replace(/\\./g,'').replace(',','.'));return (s.indexOf('▼')>=0?-1:1)*(isNaN(n)?0:n)};return p(a)-p(b)}",
                    }
                )
                html_col_indices.append(len(col_defs) - 1)
            col_defs.append(
                {
                    "headerName": "Print",
                    "field": "Print",
                    "minWidth": 90,
                    "flex": 1,
                    "sortable": True,
                    "filter": "agNumberColumnFilter",
                    "floatingFilter": True,
                    ":comparator": "(a,b)=>{const na=Number(a?a.split('.').join(''):0);const nb=Number(b?b.split('.').join(''):0);return na-nb}",
                }
            )
            if cmv:
                col_defs.append(
                    {
                        "headerName": "Δ Print",
                        "field": "Delta Print",
                        "minWidth": 100,
                        "flex": 1,
                        "sortable": True,
                        ":comparator": "(a,b)=>{const p=v=>{if(v==null)return 0;const s=String(v).replace(/<[^>]*>/g,'');const n=parseFloat(s.replace(/[^0-9.,]/g,'').replace(/\\./g,'').replace(',','.'));return (s.indexOf('▼')>=0?-1:1)*(isNaN(n)?0:n)};return p(a)-p(b)}",
                    }
                )
                html_col_indices.append(len(col_defs) - 1)
            col_defs.append(
                {
                    "headerName": "Conv",
                    "field": "Conv",
                    "minWidth": 100,
                    "flex": 1,
                    "sortable": True,
                    "filter": "agNumberColumnFilter",
                    "floatingFilter": True,
                    ":comparator": "(a,b)=>{const na=Number(a?a.replace('%','').replace(',','.'):0);const nb=Number(b?b.replace('%','').replace(',','.'):0);return na-nb}",
                }
            )
            if cmv:
                col_defs.append(
                    {
                        "headerName": "Δ Conv",
                        "field": "Delta Conv",
                        "minWidth": 100,
                        "flex": 1,
                        "sortable": True,
                        ":comparator": "(a,b)=>{const p=v=>{if(v==null)return 0;const s=String(v).replace(/<[^>]*>/g,'');const n=parseFloat(s.replace(/[^0-9.,]/g,'').replace(/\\./g,'').replace(',','.'));return (s.indexOf('▼')>=0?-1:1)*(isNaN(n)?0:n)};return p(a)-p(b)}",
                    }
                )
                html_col_indices.append(len(col_defs) - 1)
            search_inp = ui.input("🔍 Cari Outlet").props("dense outlined dark").classes("w-72 mb-3")
            grid = ui.aggrid(
                {
                    "columnDefs": col_defs,
                    "rowData": rows,
                    "pagination": True,
                    "paginationPageSize": 50,
                    "paginationPageSizeSelector": [25, 50, 100],
                    "domLayout": "autoHeight",
                    "defaultColDef": {"resizable": True},
                    "animateRows": True,
                    "rowHeight": 40,
                    "headerHeight": 40,
                    "enableCellTextSelection": True,
                },
                theme="balham",
                html_columns=html_col_indices,
            ).classes("w-full ag-theme-balham-dark").style(
                "height: auto; min-height: 300px;"
            )

            def _on_search(e):
                grid.run_grid_method("setQuickFilter", e.value)

            search_inp.on_value_change(_on_search)
        else:
            ui.label("Tidak ada outlet.").classes("text-gray-400 italic")

        if md.empty:
            return
        with ui.row().classes("w-full gap-4 mb-6"):
            with ui.card().classes("flex-[2]").style(CARD):
                ui.label("💹 Revenue by Outlet").style(ST)
                if not md.empty and "total_revenue" in md.columns:
                    tp = md.nlargest(10, "total_revenue")
                    ui.echart(
                        {
                            "tooltip": {
                                "trigger": "axis",
                                "formatter": "{b}<br/>Revenue:Rp{c}",
                            },
                            "grid": {
                                "left": "3%",
                                "right": "4%",
                                "bottom": "15%",
                                "containLabel": True,
                            },
                            "xAxis": {
                                "type": "category",
                                "data": tp["outlet_name"].tolist(),
                                "axisLabel": {
                                    "rotate": 35,
                                    "fontSize": 10,
                                    "color": "#a6adc8",
                                },
                                "axisLine": {"lineStyle": {"color": "#45475a"}},
                            },
                            "yAxis": {
                                "type": "value",
                                "axisLabel": {
                                    "color": "#a6adc8",
                                    "formatter": "Rp{value}",
                                },
                                "splitLine": {"lineStyle": {"color": "#313244"}},
                            },
                            "series": [
                                {
                                    "name": "Revenue",
                                    "type": "bar",
                                    "data": tp["total_revenue"].tolist(),
                                    "itemStyle": {
                                        "color": {
                                            "type": "linear",
                                            "x": 0,
                                            "y": 0,
                                            "x2": 0,
                                            "y2": 1,
                                            "colorStops": [
                                                {"offset": 0, "color": "#89b4fa"},
                                                {"offset": 1, "color": "#45475a"},
                                            ],
                                        }
                                    },
                                    "barMaxWidth": 40,
                                    "label": {
                                        "show": True,
                                        "position": "top",
                                        "formatter": "{@value}",
                                        "color": "#cdd6f4",
                                        "fontSize": 10,
                                    },
                                }
                            ],
                            **ECHART,
                        }
                    ).classes("w-full h-[350px]")
                else:
                    ui.label("Tidak ada data.").classes("text-gray-400 italic")
            with ui.card().classes("flex-1").style(CARD):
                ui.label("🔄 Conversion Funnel").style(ST)
                ft = int(md["foto_qty"].sum()) if "foto_qty" in md.columns else 0
                ul = int(md["unlock_qty"].sum()) if "unlock_qty" in md.columns else 0
                pr = int(md["print_qty"].sum()) if "print_qty" in md.columns else 0
                ui.echart(
                    {
                        "tooltip": {"trigger": "item", "formatter": "{b}:{c}"},
                        "series": [
                            {
                                "type": "funnel",
                                "left": "10%",
                                "top": 20,
                                "bottom": 20,
                                "width": "80%",
                                "min": 0,
                                "max": ft,
                                "minSize": "0%",
                                "maxSize": "100%",
                                "sort": "descending",
                                "gap": 2,
                                "label": {
                                    "show": True,
                                    "position": "inside",
                                    "color": "#fff",
                                    "fontSize": 12,
                                    "formatter": "{b}:{c}",
                                },
                                "itemStyle": {
                                    "borderColor": "#1e1e2e",
                                    "borderWidth": 2,
                                },
                                "data": [
                                    {
                                        "value": ft,
                                        "name": "📸 Foto",
                                        "itemStyle": {"color": "#89b4fa"},
                                    },
                                    {
                                        "value": ul,
                                        "name": "🔓 Unlock",
                                        "itemStyle": {"color": "#f9e2af"},
                                    },
                                    {
                                        "value": pr,
                                        "name": "🖨️ Print",
                                        "itemStyle": {"color": "#a6e3a1"},
                                    },
                                ],
                            }
                        ],
                        **{
                            k: v
                            for k, v in ECHART.items()
                            if k not in ("xAxis", "yAxis")
                        },
                    }
                ).classes("w-full h-[300px]")
        ui.separator().classes("my-6")
        # HANYA tabel Tren Omset: tampilkan outlet dengan transaksi terakhir <= 30 hari
        last_tx = _load_last_tx_dates()
        if last_tx:
            cutoff = (datetime.now() - timedelta(days=30)).strftime("%Y-%m-%d")
            ols = sorted(o for o, d in last_tx.items() if d >= cutoff)
        else:
            ols = (
                md["outlet_name"].dropna().unique().tolist()
                if "outlet_name" in md.columns
                else []
            )
        if ols:
            tr = a.build_trend_table(_ff, _cp, ols, 12)
            if tr["has_data"]:
                vc = tr["value_cols"]
                ad = tr["active_df"]
                idf = tr["inactive_df"]
                with ui.expansion(
                    "📆 Tren Omset Outlet (12 Bulan)", icon="trending_up"
                ).classes("w-full mb-4"):
                    if not ad.empty:
                        ui.label("Outlet Aktif").classes(
                            "text-sm font-semibold text-green-400 mt-2 mb-2"
                        )
                        rr = []
                        for _, rd in ad.iterrows():
                            d = {
                                "Outlet": str(rd.get("Outlet", "")),
                                "Rata-rata": str(rd.get("Rata-rata", "Rp 0")),
                            }
                            for p in vc:
                                d[p] = str(rd.get(p, "Rp 0"))
                            rr.append(d)
                        tr_cols = [
                            {
                                "name": "Outlet",
                                "label": "Outlet",
                                "field": "Outlet",
                                "align": "left",
                                "sortable": True,
                            }
                        ]
                        tr_cols.append(
                            {
                                "name": "Rata-rata",
                                "label": "Rata-rata",
                                "field": "Rata-rata",
                                "align": "right",
                                "sortable": True,
                            }
                        )
                        for p in vc:
                            tr_cols.append(
                                {"name": p, "label": p, "field": p, "align": "right"}
                            )
                        # Color each month value green/red vs prev month
                        trend_html_indices = list(range(2, 2 + len(vc)))
                        for ri in rr:
                            months = sorted(list(vc))  # urut kronologis: terlama -> terbaru
                            prev_val = None
                            for p in months:
                                raw = ri.get(p, "Rp 0")
                                try:
                                    s = raw.replace("Rp ", "").strip()
                                    cur = float(s.replace(".", "").replace(",", ""))
                                except:
                                    cur = 0
                                # Hijau jika omset > bulan sebelumnya, merah jika < bulan sebelumnya
                                if prev_val is not None and cur != prev_val:
                                    if cur > prev_val:
                                        ri[p] = (
                                            '<span style="color:#a6e3a1;font-weight:600">'
                                            + raw
                                            + "</span>"
                                        )
                                    else:
                                        ri[p] = (
                                            '<span style="color:#f38ba8;font-weight:600">'
                                            + raw
                                            + "</span>"
                                        )
                                prev_val = cur
                        async def _download_active_excel():
                            selected = await active_grid.get_selected_rows() if active_grid else []
                            _download_trend_excel(selected, vc, "Outlet Aktif")

                        async def _download_active_pdf():
                            selected = await active_grid.get_selected_rows() if active_grid else []
                            _download_trend_pdf(selected, vc, "Outlet Aktif")

                        with ui.row().classes("items-center gap-2 mb-2"):
                            ui.button("⬇️ Excel", on_click=_download_active_excel).props("dense color=green")
                            ui.button("⬇️ PDF", on_click=_download_active_pdf).props("dense color=red")
                            ui.label("Centang outlet di tabel, lalu pilih format download.").classes("text-xs text-gray-500")

                        active_grid = ui.aggrid(
                            {
                                "columnDefs": [
                                    {
                                        "headerName": "Outlet",
                                        "field": "Outlet",
                                        "pinned": "left",
                                        "minWidth": 180,
                                        "flex": 2,
                                        "sortable": True,
                                        "filter": "agTextColumnFilter",
                                        "floatingFilter": True,
                                        "checkboxSelection": True,
                                        "headerCheckboxSelection": True,
                                    },
                                    {
                                        "headerName": "Rata-rata",
                                        "field": "Rata-rata",
                                        "minWidth": 130,
                                        "flex": 1,
                                        "sortable": True,
                                        "filter": "agNumberColumnFilter",
                                        "floatingFilter": True,
                                        ":comparator": "(a,b)=>{const x=a.split(' ').slice(-1)[0];const y=b.split(' ').slice(-1)[0];return Number(x?x.split('.').join(''):0)-Number(y?y.split('.').join(''):0)}",
                                    },
                                ]
                                + [
                                    {
                                        "headerName": p,
                                        "field": p,
                                        "minWidth": 110,
                                        "flex": 1,
                                        "sortable": True,
                                        ":comparator": "(a,b)=>{const p=v=>{if(v==null)return 0;const s=String(v).replace(/<[^>]*>/g,'').replace('Rp ','').replace(/\\./g,'');return parseFloat(s)||0};return p(a)-p(b)}",
                                    }
                                    for p in vc
                                ],
                                "rowData": rr,
                                "pagination": True,
                                "paginationPageSize": 50,
                                "paginationPageSizeSelector": [25, 50, 100],
                                "domLayout": "autoHeight",
                                "defaultColDef": {"resizable": True},
                                "animateRows": True,
                                "rowHeight": 40,
                                "headerHeight": 40,
                                "enableCellTextSelection": True,
                                "rowSelection": "multiple",
                                "suppressRowClickSelection": True,
                            },
                            theme="balham",
                            html_columns=trend_html_indices,
                        ).classes("w-full ag-theme-balham-dark").style(
                            "height: auto; min-height: 200px;"
                        )
                    else:
                        ui.label("Tidak ada outlet aktif.").classes(
                            "text-gray-400 italic"
                        )
                    if not idf.empty:
                        ui.label("Outlet Tidak Aktif").classes(
                            "text-sm font-semibold text-gray-400 mt-2 mb-2"
                        )
                        rr = []
                        for _, rd in idf.iterrows():
                            d = {
                                "Outlet": str(rd.get("Outlet", "")),
                                "Rata-rata": str(rd.get("Rata-rata", "Rp 0")),
                            }
                            for p in vc:
                                d[p] = str(rd.get(p, "Rp 0"))
                            rr.append(d)
                        trend_html_indices2 = list(range(2, 2 + len(vc)))
                        for ri in rr:
                            months = sorted(list(vc))  # urut kronologis: terlama -> terbaru
                            prev_val = None
                            for p in months:
                                raw = ri.get(p, "Rp 0")
                                try:
                                    s = raw.replace("Rp ", "").strip()
                                    cur = float(s.replace(".", "").replace(",", ""))
                                except:
                                    cur = 0
                                # Hijau jika omset > bulan sebelumnya, merah jika < bulan sebelumnya
                                if prev_val is not None and cur != prev_val:
                                    if cur > prev_val:
                                        ri[p] = (
                                            '<span style="color:#a6e3a1;font-weight:600">'
                                            + raw
                                            + "</span>"
                                        )
                                    else:
                                        ri[p] = (
                                            '<span style="color:#f38ba8;font-weight:600">'
                                            + raw
                                            + "</span>"
                                        )
                                prev_val = cur
                        async def _download_inactive_excel():
                            selected = await inactive_grid.get_selected_rows() if inactive_grid else []
                            _download_trend_excel(selected, vc, "Outlet Tidak Aktif")

                        async def _download_inactive_pdf():
                            selected = await inactive_grid.get_selected_rows() if inactive_grid else []
                            _download_trend_pdf(selected, vc, "Outlet Tidak Aktif")

                        with ui.row().classes("items-center gap-2 mb-2"):
                            ui.button("⬇️ Excel", on_click=_download_inactive_excel).props("dense color=green")
                            ui.button("⬇️ PDF", on_click=_download_inactive_pdf).props("dense color=red")
                            ui.label("Centang outlet di tabel, lalu pilih format download.").classes("text-xs text-gray-500")

                        inactive_grid = ui.aggrid(
                            {
                                "columnDefs": [
                                    {
                                        "headerName": "Outlet",
                                        "field": "Outlet",
                                        "pinned": "left",
                                        "minWidth": 180,
                                        "flex": 2,
                                        "sortable": True,
                                        "filter": "agTextColumnFilter",
                                        "floatingFilter": True,
                                        "checkboxSelection": True,
                                        "headerCheckboxSelection": True,
                                    },
                                    {
                                        "headerName": "Rata-rata",
                                        "field": "Rata-rata",
                                        "minWidth": 130,
                                        "flex": 1,
                                        "sortable": True,
                                        "filter": "agNumberColumnFilter",
                                        "floatingFilter": True,
                                        ":comparator": "(a,b)=>{const x=a.split(' ').slice(-1)[0];const y=b.split(' ').slice(-1)[0];return Number(x?x.split('.').join(''):0)-Number(y?y.split('.').join(''):0)}",
                                    },
                                ]
                                + [
                                    {
                                        "headerName": p,
                                        "field": p,
                                        "minWidth": 110,
                                        "flex": 1,
                                        "sortable": True,
                                        ":comparator": "(a,b)=>{const p=v=>{if(v==null)return 0;const s=String(v).replace(/<[^>]*>/g,'').replace('Rp ','').replace(/\\./g,'');return parseFloat(s)||0};return p(a)-p(b)}",
                                    }
                                    for p in vc
                                ],
                                "rowData": rr,
                                "pagination": True,
                                "paginationPageSize": 50,
                                "paginationPageSizeSelector": [25, 50, 100],
                                "domLayout": "autoHeight",
                                "defaultColDef": {"resizable": True},
                                "animateRows": True,
                                "rowHeight": 40,
                                "headerHeight": 40,
                                "enableCellTextSelection": True,
                                "rowSelection": "multiple",
                                "suppressRowClickSelection": True,
                            },
                            theme="balham",
                            html_columns=trend_html_indices2,
                        ).classes("w-full ag-theme-balham-dark").style(
                            "height: auto; min-height: 200px;"
                        )
                    ui.label(
                        "Nilai kosong = 0. Rata-rata dari 12 bulan (bulan berjalan tidak dihitung), hanya omset > 0 yang dihitung."
                    ).classes("text-[10px] text-gray-500 italic")

            # 📆 Tren Omset Outlet (30 Hari) — harian, 30 hari terakhir berakhir H-1
            try:
                daily_data = _load_daily_data()
                if daily_data:
                    ddf = pd.DataFrame(daily_data)
                    dtr = a.build_daily_trend_table(ddf, ols, days=30)
                    if dtr["has_data"]:
                        ddays = dtr["value_cols"]
                        with ui.expansion(
                            "📆 Tren Omset Outlet (30 Hari)", icon="trending_up"
                        ).classes("w-full mb-4"):
                            drr = []
                            for _, rd in dtr["df"].iterrows():
                                d = {
                                    "Outlet": str(rd.get("Outlet", "")),
                                    "Total 30 Hari": str(rd.get("Total 30 Hari", "Rp 0")),
                                }
                                for p in ddays:
                                    d[p] = str(rd.get(p, "Rp 0"))
                                drr.append(d)
                            d_html = list(range(2, 2 + len(ddays)))
                            # warna: hijau jika omset > hari sebelumnya, merah jika < (urutan kronologis)
                            for ri in drr:
                                days_asc = sorted(list(ddays))
                                prev_val = None
                                for p in days_asc:
                                    raw = ri.get(p, "Rp 0")
                                    try:
                                        s = raw.replace("Rp ", "").strip()
                                        cur = float(s.replace(".", "").replace(",", ""))
                                    except Exception:
                                        cur = 0
                                    if prev_val is not None and cur != prev_val:
                                        if cur > prev_val:
                                            ri[p] = (
                                                '<span style="color:#a6e3a1;font-weight:600">'
                                                + raw
                                                + "</span>"
                                            )
                                        else:
                                            ri[p] = (
                                                '<span style="color:#f38ba8;font-weight:600">'
                                                + raw
                                                + "</span>"
                                            )
                                    prev_val = cur
                            dcols = [
                                {
                                    "headerName": "Outlet",
                                    "field": "Outlet",
                                    "pinned": "left",
                                    "minWidth": 180,
                                    "flex": 2,
                                    "sortable": True,
                                    "filter": "agTextColumnFilter",
                                    "floatingFilter": True,
                                },
                                {
                                    "headerName": "Total 30 Hari",
                                    "field": "Total 30 Hari",
                                    "minWidth": 130,
                                    "flex": 1,
                                    "sortable": True,
                                    "filter": "agNumberColumnFilter",
                                    "floatingFilter": True,
                                    ":comparator": "(a,b)=>{const x=a.split(' ').slice(-1)[0];const y=b.split(' ').slice(-1)[0];return Number(x?x.split('.').join(''):0)-Number(y?y.split('.').join(''):0)}",
                                },
                            ] + [
                                {
                                    "headerName": p,
                                    "field": p,
                                    "minWidth": 100,
                                    "flex": 1,
                                    "sortable": True,
                                    ":comparator": "(a,b)=>{const p=v=>{if(v==null)return 0;const s=String(v).replace(/<[^>]*>/g,'').replace('Rp ','').replace(/\\./g,'');return parseFloat(s)||0};return p(a)-p(b)}",
                                }
                                for p in ddays
                            ]
                            ui.aggrid(
                                {
                                    "columnDefs": dcols,
                                    "rowData": drr,
                                    "pagination": True,
                                    "paginationPageSize": 50,
                                    "paginationPageSizeSelector": [25, 50, 100],
                                    "domLayout": "autoHeight",
                                    "defaultColDef": {"resizable": True},
                                    "animateRows": True,
                                    "rowHeight": 40,
                                    "headerHeight": 40,
                                    "enableCellTextSelection": True,
                                },
                                theme="balham",
                                html_columns=d_html,
                            ).classes("w-full ag-theme-balham-dark").style(
                                "height: auto; min-height: 200px;"
                            )
                            ui.label(
                                "Warna: hijau = omset lebih tinggi dari hari sebelumnya, merah = lebih rendah."
                            ).classes("text-[10px] text-gray-500 italic")
            except Exception:
                pass
        else:
            ui.label("Tidak ada outlet.").classes("text-gray-400 italic")


def set_filters(df, fdf, cp, cmp, area="Semua", kat="Semua", tip="Semua"):
    global _df, _ff, _cp, _cmp
    _df = df
    _ff = fdf
    _cp = cp
    _cmp = cmp
    if _act_sel is not None:
        periods = _periods
        _act_sel.options = periods
        _act_sel.value = cp or periods[-1]
        cmp_opts = ["-"] + [p for p in periods if p != (cp or periods[-1])]
        _cmp_sel.options = cmp_opts
        _cmp_sel.value = cmp if cmp in cmp_opts else "-"
    if _content is not None:
        _build_outlet_table(_cp, _cmp)


def create_page(c):
    global _content, _act_sel, _cmp_sel, _periods, _a
    _a = get_adapter()
    _content = ui.column().classes("w-full")
    with c:
        ui.add_head_html(TBL_CSS)
        with ui.row().classes("w-full items-center gap-3 mb-2"):
            ui.label("📸").classes("text-3xl")
            ui.label("difotoin.id").classes("text-2xl font-bold text-white")
            ui.label("— Dashboard").classes("text-lg text-gray-400")
            ui.label(f"• {datetime.now().strftime('%d %b %Y %H:%M')}").classes(
                "text-xs text-gray-500 ml-auto"
            )
        ui.separator().classes("mb-3")

        # ── Last Update Info ──
        _info = _get_last_update_info()
        with ui.row().classes("w-full items-center gap-2 mb-3"):
            cache_str = _info.get("cache_time") or "-"
            date_str = _info.get("latest_date") or "-"
            hours = _info.get("hours_old")
            if hours is not None and hours > 24:
                color = "#ef4444"
                icon = "🔴"
            elif hours is not None and hours > 8:
                color = "#f59e0b"
                icon = "🟡"
            else:
                color = "#22c55e"
                icon = "🟢"
            ui.label(f"{icon} Update cache: {cache_str}").classes("text-xs").style(f"color: {color}")
            ui.label(f"• Data terakhir: {date_str}").classes("text-xs text-gray-400")
            if hours is not None:
                ui.label(f"({hours:.0f}h yang lalu)").classes("text-xs").style(f"color: {color}")
            else:
                ui.label("(belum ada sync)").classes("text-xs text-red-400")
            ui.label("• Sync tiap 1 jam").classes("text-xs text-gray-500 ml-auto")
        ui.separator().classes("mb-6")

        # STABLE dropdowns (never destroyed)
        data = _ff if _ff is not None else _a.load_full_data()
        if data is not None and not data.empty:
            _periods = (
                sorted(data["periode"].dropna().astype(str).unique().tolist())
                if "periode" in data.columns
                else []
            )
        if not _periods:
            _periods = ["-"]
        with ui.row().classes("w-full items-center gap-4 mb-3 flex-wrap"):
            ui.label("Periode:").classes("text-xs text-gray-400")
            _act_sel = (
                ui.select(_periods, value=_cp or _periods[-1], label="Bulan Aktif")
                .props("dense outlined dark")
                .classes("w-40")
            )
            ui.label("Bandingkan:").classes("text-xs text-gray-400")
            _cmp_sel = (
                ui.select(["-"], value="-", label="Bandingkan")
                .props("dense outlined dark")
                .classes("w-40")
            )

        # Init compare options
        if _act_sel.value:
            _cmp_sel.options = ["-"] + [p for p in _periods if p != _act_sel.value]
            _cmp_sel.value = _cmp if _cmp in _cmp_sel.options else "-"

        # Wire callbacks
        def _on_change():
            global _cp, _cmp
            _cp = _act_sel.value
            _cmp = None if _cmp_sel.value == "-" else _cmp_sel.value
            _build_outlet_table(_cp, _cmp)

        _act_sel.on("update:model-value", _on_change)
        _cmp_sel.on("update:model-value", _on_change)

        # Build initial content
        _build_outlet_table(_cp or (_periods[-1] if _periods else None), _cmp)
